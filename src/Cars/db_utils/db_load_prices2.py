'''
Created on May 31, 2023

@author: DNP Enterprises Inc.
Loads the cars data into a pandas dataframe  from spreadsheets
'''
import openpyxl
import pandas as pd

if __name__ == '__main__':
    car_data_file = 'C:/Users/dpenn/Desktop/Cars/CarData.xlsx'
    dealer_wkbk = openpyxl.load_workbook(car_data_file)
    prices = pd.DataFrame(columns = ['ID', 'Year', 'Make', 'Model', 'Price', 'Stock #', 'URL']) #create dataframe ready for loading
    total_record_count = 0    
    for index, dealer_sheet in enumerate(dealer_wkbk.sheetnames): #get each dealer sheet
        print (dealer_sheet)
        dealer_wksh = dealer_wkbk[dealer_sheet]
        for index, row in enumerate(dealer_wksh.rows): # for each sheet, scan the rows for the dealer price sheet file name
            if index == 0:
                continue # skip the first row
            dealer_price_file = row[6].value
            print (dealer_price_file)
            print ("------------------------------------------------------------------------------------------------------------------------")
            price_wkbk = openpyxl.load_workbook(dealer_price_file)
            dealer_record_count = 0
            price_wksht = price_wkbk.active
            for row in price_wksht.rows:
                dealer_id = row[0].value
                year = row[1].value
                make = row[2].value
                model = row[3].value
                price = row[4].value
                stock_num =row[5].value
                link = row[6].value
                print ("ID: ", dealer_id, "Year: ", year, "Make: ", make, "Model: ", model, "Price: ", price, "Stock #: ", stock_num, "URL: ", link)
                val = pd.DataFrame([{'ID': dealer_id, 'Year': year, 'Make': make, 'Model': model, 'Price': price, 'Stock #': stock_num, 'URL':link}])
                prices = pd.concat([prices, val])
                dealer_record_count +=1
            print ("Dealer record count: ", dealer_record_count)
            total_record_count = total_record_count + dealer_record_count
    print ("Total records: " , total_record_count)
    #prices.to_excel('C:/Users/dpenn/Desktop/Cars/prices.xlsx', index = False)
    prices.to_csv('C:/Users/dpenn/Desktop/Cars/prices.csv', index = False)
